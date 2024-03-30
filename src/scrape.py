import csv
import pandas as pd
import re
import numpy as np
import os
import logging
import requests
import fitz 
from bs4 import BeautifulSoup
from io import BytesIO
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


class WebScraper:
    """
    A base class for web scraping that manages session creation and applies filters for data collection.

    Attributes:
        session (requests.Session): A session for making HTTP requests.
        rest (int): Default time to wait for page elements to load.
        filter_options (dict): Specifies the periods in a list for which to collect data ['AM', 'PM', 'MD'].
        By default all three periods will be scraped. 

    Methods:
        __init__(self, period_filters=None): Initializes the WebScraper with optional period filters.
    """
    def __init__(self, period_filters = None):
        """
        Initializes the WebScraper instance with specified period filters.

        Args:
            period_filters (dict, optional): Filters for data collection periods. Defaults to None.
        """
        self.session = requests.Session()
        self.rest = 10
        self.filter_options = {
            'AM': True,
            'PM': True,
            'MD': True
        }
        if period_filters:
            self._set_filters(period_filters)
        self._apply_filters()

    def _set_filters(self, period_filters):
        for filter_key in period_filters:
            if filter_key in self.filter_options:
                self.filter_options[filter_key] = True

    def _apply_filters(self):
        self.collect_am = self.filter_options.get('AM')
        self.collect_pm = self.filter_options.get('PM')
        self.collect_md = self.filter_options.get('MD')
        
class OldTMCScraper(WebScraper):
    """
    Scraper for old Traffic Movement Count (TMC) data (before 2012), inheriting from WebScraper.

    Methods:
        __init__(self, raw_directory, period_filters=None): Initializes the scraper with a directory and optional filters.
        scrape_sub_urls(self): Expands and scrapes sub URLs from the raw data files.
        scrape_tmc_counts(self, sub_urls_df): Scrapes TMC data from expanded sub URLs.
    """
    def __init__(self, raw_directory, period_filters = None):
        """
        Initializes the OldTMCScraper with a directory for raw data and optional period filters.

        Args:
            raw_directory (str): The directory containing raw data files.
            period_filters (dict, optional): Filters for data collection periods. Defaults to None.
        """
        super().__init__(period_filters=period_filters) 
        self.raw = raw_directory
        self.raw_df = self._load_csv()        
        self.options = Options()
        self.options.add_argument('--headless')
        self.driver = webdriver.Chrome(options=self.options)
        
    def _load_csv(self):
        return pd.read_csv(self.raw,sep=';')
    
    def scrape_sub_urls(self):
        """
        Scrapes and expands URLs from the raw data file.

        Returns:
            DataFrame: A pandas DataFrame containing expanded sub URLs.
        """
        urls = self.raw_df["URL"] 
        expanded_urls= urls.apply(lambda row: self._explode_url(row))
        expanded_urls.rename('Sub_URL',inplace=True)
        return pd.concat([self.raw_df,expanded_urls],axis =1).explode\
                        ('Sub_URL',ignore_index=True)
    
    def _explode_url(self, row):
        
        if row:
            sub_urls = []
            self.driver.get('about:blank') 
            self.driver.get(row)

            try:
                coordinate_element = WebDriverWait(self.driver, self.rest).until(
                    EC.presence_of_element_located((By.CLASS_NAME, 'sorting_1'))
                )
                coordinate = coordinate_element.text.strip()

                links_selector = f'td a[target^="0{coordinate}"]'

                links_elements = WebDriverWait(self.driver, self.rest).until(
                    EC.presence_of_all_elements_located((By.CSS_SELECTOR, links_selector))
                )

                for link in links_elements:
                    href = link.get_attribute('href')
                    sub_urls.append(href)

            except Exception as e:
                print("An error occurred: ", e)
            
            return sub_urls
    
    def scrape_tmc_counts(self,sub_urls_df):
        """
        Scrapes TMC data from the provided DataFrame of sub URLs.

        Args:
            sub_urls_df (DataFrame): A DataFrame containing sub URLs of .htm files.

        Returns:
            DataFrame: A normalized DataFrame with scraped TMC data.
        """
        tmc_urls = sub_urls_df['Sub_URL'] 
        tmc_data = tmc_urls.apply(lambda row: self._scrape_tmc(row))
        return pd.json_normalize(tmc_data)

    def _scrape_tmc(self, row):
        if pd.isna(row):
            return None

        try:
            response = self.session.get(row)
            soup = BeautifulSoup(response.text, 'html.parser')
            tmc = {}
            available_am = False
            available_md = False
            available_pm = False
            control_dict = {
                'AM':{'check':None,'table_key':None},
                'MD':{'check':None,'table_key':None},
                'PM':{'check':None,'table_key':None}
                }
            for i in range(2,5):
                period = self._get_peak_hour(soup,key=i)
                if period:
                    peak_start_hr = int(period[:2])

                    if 6 <= peak_start_hr <= 8:
                        control_dict['AM']['table_key'] = i
                        available_am = True
                    elif 9 <= peak_start_hr <= 15:
                        control_dict['MD']['table_key'] = i
                        available_md = True
                    elif 16 <= peak_start_hr <= 19:
                        control_dict['PM']['table_key'] = i
                        available_pm = True
                    else:
                        continue
            
            control_dict['AM']['check'] = self.collect_am and available_am
            control_dict['MD']['check'] = self.collect_md and available_md
            control_dict['PM']['check'] = self.collect_pm and available_pm
            
            tmc.update({
                'date': self._get_date(soup),
                'weather': self._get_weather(soup),
                'type': self._get_type(soup),
                'AM_available': available_am,
                'AM_scraped': self.collect_am,
                'MD_available': available_md,
                'MD_scraped': self.collect_md,                
                'PM_available': available_pm,
                'PM_scraped': self.collect_pm
            })
                
            for key,value in control_dict.items():
                if value.get('check'):
                    tag = key
                    position = value.get('table_key')
                    tmc.update({
                        f'{tag}_peak_hour': self._get_peak_hour(soup,position),
                        f'{tag}_north_bikes_vol': self._get_north_bikes_vol(soup,position),
                        f'{tag}_north_peds_vol': self._get_north_peds_vol(soup,position),
                        f'{tag}_north_veh_vol': self._get_north_veh_vol(soup,position),
                        f'{tag}_east_bikes_vol' : self._get_east_bikes_vol(soup,position),
                        f'{tag}_east_peds_vol' : self._get_east_peds_vol(soup,position),
                        f'{tag}_east_veh_vol' : self._get_east_veh_vol(soup,position),
                        f'{tag}_south_bikes_vol' : self._get_south_bikes_vol(soup,position),
                        f'{tag}_south_peds_vol' : self._get_south_peds_vol(soup,position),
                        f'{tag}_south_veh_vol' : self._get_south_veh_vol(soup,position),
                        f'{tag}_west_bikes_vol' : self._get_west_bikes_vol(soup,position),
                        f'{tag}_west_peds_vol' : self._get_west_peds_vol(soup,position),
                        f'{tag}_west_veh_vol' : self._get_west_veh_vol(soup,position)
                    })
            
            return tmc
            
        except Exception as e:
            logging.error(f'{row} had an error: {e}')
            return None

    def _get_date(self, soup):
        element = soup.select_one("th[valign='TOP']:nth-of-type(3) p")
        return pd.to_datetime(element.text if element else None)

    def _get_weather(self, soup):
        elements = soup.select("th:nth-of-type(13) p")
        return elements[0].text.split()[-1].lower() if elements else None

    def _get_type(self, soup):
        element = soup.select_one("th:nth-of-type(8) p")
        return element.text if element else None

    def _get_peak_hour(self, soup, key):
        element = soup.select_one(f"table:nth-of-type({key}) [valign='MIDDLE'] p.p8")
        return element.text.replace("Maximum Hour ", "") if element else None

    def _get_north_bikes_vol(self, soup, key):
        elements = soup.select(f"table:nth-of-type({key}) [valign='BOTTOM'] tr:-soup-contains('Bikes') tr:nth-of-type(1)")
        return re.findall(r'\d+', elements[0].text)[0] if elements else None

    def _get_north_peds_vol(self, soup, key):
        elements = soup.select(f"table:nth-of-type({key}) table:nth-of-type(1) th:-soup-contains('PEDs')")
        return re.findall(r'\d+', elements[0].text)[-1] if elements else None

    def _get_north_veh_vol(self, soup, key):
        elements = soup.select(f"table:nth-of-type({key}) table:nth-of-type(1) th.s")
        nums = re.findall(r'\d+', "".join([str(e) for e in elements]))
        return nums[2] if len(nums) > 2 else None
    
    def _get_east_bikes_vol(self, soup, key):
        elements = soup.select("tr:-soup-contains('Bikes') th.s:-soup-contains('Bikes')")
        return re.findall(r'\d+', elements[2*key-3].text)[0] if elements and 0 <= 2*key-3 < len(elements) else None
    
    def _get_east_peds_vol(self, soup, key):
        elements = soup.select("tr:-soup-contains('Peds') th.s:-soup-contains('Peds')")
        return re.findall(r'\d+', elements[2*key-3].text)[0] if elements and 0 <= 2*key-3 < len(elements) else None

    def _get_east_veh_vol(self, soup, key):
        elements = soup.select("tr:-soup-contains('Peds') th[align='LEFT'] th:nth-of-type(2) th.s:not(:-soup-contains('Peds')):not(:-soup-contains('Bikes'))")
        return re.findall(r'\d+', elements[key-2].text)[0] if elements else None
    
    def _get_south_bikes_vol(self, soup, key):
        elements = soup.select("tr:nth-of-type(5):-soup-contains('Bikes')")
        return re.findall(r'\d+', elements[key-2].text)[0] if elements and 0 <= key-2 < len(elements) else None

    def _get_south_peds_vol(self, soup, key):
        elements = soup.select("tr:nth-of-type(4) th.s:-soup-contains('PEDs')")
        return re.findall(r'\d+', elements[key-2].text)[0] if elements and 0 <= key-2 < len(elements) else None
 
    def _get_south_veh_vol(self, soup, key):
        elements = soup.select(f"tr:nth-of-type({5*key-6}) th.s:not(:-soup-contains('PEDs'))")
        return re.findall(r'\d+', elements[-1].text)[0] if elements else None
    
    def _get_west_bikes_vol(self, soup, key):
        elements = soup.select("tr:-soup-contains('Bikes') th.s:-soup-contains('Bikes')")
        return re.findall(r'\d+', elements[2*key-4].text)[0] if elements and 0 <= 2*key-4 < len(elements) else None

    def _get_west_peds_vol(self, soup, key):
        elements = soup.select("tr:-soup-contains('Peds') th.s:-soup-contains('Peds')")
        return re.findall(r'\d+', elements[2*key-4].text)[0] if elements and 0 <= 2*key-4 < len(elements) else None

    def _get_west_veh_vol(self, soup, key):
        elements = soup.select("tr:-soup-contains('Peds') th[align='RIGHT'] th:nth-of-type(3) th.s:not(:-soup-contains('Peds')):not(:-soup-contains('Bikes'))")
        return re.findall(r'\d+', elements[key-2].text)[0] if elements else None

class RecentTMCScraper(WebScraper):
    """
    Scraper for recent Traffic Moevement Count (TMC) data (after 2012), inheriting from WebScraper.

    Methods:
        __init__(self, raw_directory, period_filters=None): Initializes the scraper with a directory and optional filters.
        scrape_sub_urls(self): Constructs URLs for recent TMC data based on year and object ID.
        scrape_tmc_counts(self, sub_urls_df): Scrapes TMC data from URLs provided in a DataFrame.
    """
    def __init__(self, raw_directory, period_filters = None):
        """
        Initializes the RecentTMCScraper with a directory for raw data and optional period filters.

        Args:
            raw_directory (str): The directory containing raw data files.
            period_filters (dict, optional): Filters for data collection periods. Defaults to None.
        """
        super().__init__(period_filters=period_filters) 
        self.raw = raw_directory
        self.raw_df = self._load_csv()  
        self.year_key_dict = {
            '2012' : '16',
            '2013' : '16',
            '2014' : '16',
            '2015' : '16',
            '2016' : '16',
            '2017' : '16',
            '2018' : '211', 
            '2019' : '23',  
            '2020' : '21',                            
            '2021' : '20',
            '2022' : '217'
        }  
        
    def _load_csv(self):
        return pd.read_csv(self.raw)
    
    def scrape_sub_urls(self):
        """
        Constructs and retrieves URLs for recent TMC data.

        Returns:
            DataFrame: A DataFrame with constructed URLs and file types of .pdf or .xlsx.
        """
        tmc = self.raw_df[["Year","OBJECTID"]].astype(str)
        constructed_data = tmc.apply(lambda row: self._construct_url(row), axis=1)
        constructed_urls = constructed_data.apply(lambda x: x[0] if x is not np.nan else np.nan)
        file_types = constructed_data.apply(lambda x: x[1] if x is not None else np.nan)
        constructed_urls.rename('Const_URL', inplace=True)
        file_types.rename('FileType', inplace=True)
        return pd.concat([self.raw_df, constructed_urls, file_types], axis=1)
    
    def _construct_url(self, row):
        year = row['Year']
        object_id = row['OBJECTID']
        base_url = "https://maps.vancouver.ca"
        path_url = f"/server/rest/services/VanMapViewer/Traffic_and_Transportation/MapServer/{self.year_key_dict.get(year)}/{object_id}/attachments/"

        response = self.session.get(base_url+path_url)
        response.raise_for_status() 
        soup_link = BeautifulSoup(response.text, 'html.parser')
        attachment_link_tag = soup_link.find('a', text=lambda x: '.xlsx' in x.lower() or '.pdf' in x.lower())

        if attachment_link_tag:
            attachment_url = attachment_link_tag['href']
            full_url = base_url + attachment_url
            file_type = 'xlsx' if '.xlsx' in attachment_link_tag.text else 'pdf'
            return full_url, file_type
        else:
            print(f'The tmc link on {year} for {object_id} was not found')
            return np.nan, None

    def scrape_tmc_counts(self,sub_urls_df):
        """
        Scrapes TMC data from the URLs provided in the DataFrame.

        Args:
            sub_urls_df (DataFrame): A DataFrame containing URLs and file types of .pdf or .xlsx.

        Returns:
            DataFrame: A normalized DataFrame with scraped TMC data.
        """
        tmc_urls = sub_urls_df[['Const_URL','FileType']]
        tmc_data = tmc_urls.apply(lambda row: self._scrape_tmc(row),axis=1)
        return pd.json_normalize(tmc_data)

    def _scrape_tmc(self, row):
        if pd.isna(row['Const_URL']):
            return None
        elif row['FileType'] == 'pdf':
            return self._read_pdf(row['Const_URL'])
        elif row['FileType'] == 'xlsx':
            return self._read_xlsx(row['Const_URL'])
        else:
            return None
        
    def _read_xlsx(self,url):
      
        try:
            response = self.session.get(url)
            response.raise_for_status()
            file_content = BytesIO(response.content)
            wb = load_workbook(filename=file_content)
            ws = wb['Summary']

            tmc = {}

            available_am = False
            available_md = False
            available_pm = False
            
            control_dict = {
                'AM':{'check':None,'table_key':None},
                'MD':{'check':None,'table_key':None},
                'PM':{'check':None,'table_key':None}
                }

            for i in [32,62,92]:

                period = self._xlsx_get_peak_hour(ws,key=i)

                if period:
                    
                    peak_start = period[:6]
                    peak_end= period[-5:]

                    period_check = (pd.to_datetime(peak_end) - pd.to_datetime(peak_start)).total_seconds() == 3600

                    if  i==32 and period_check:
                        control_dict['AM']['table_key'] = i
                        available_am = True
                    elif i==62 and period_check:
                        control_dict['MD']['table_key'] = i
                        available_md = True
                    elif i==92 and period_check:
                        control_dict['PM']['table_key'] = i
                        available_pm = True
                    else:
                        continue
            
            control_dict['AM']['check'] = self.collect_am and available_am
            control_dict['MD']['check'] = self.collect_md and available_md
            control_dict['PM']['check'] = self.collect_pm and available_pm

            tmc.update({
                'date': self._xlsx_get_date(ws),
                'weather': self._xlsx_get_weather(ws),
                'type': self._xlsx_get_type(ws),
                'AM_available': available_am,
                'AM_scraped': self.collect_am,
                'MD_available': available_md,
                'MD_scraped': self.collect_md,                
                'PM_available': available_pm,
                'PM_scraped': self.collect_pm
            })
                
            for key,value in control_dict.items():
                if value.get('check'):
                    tag = key
                    position = value.get('table_key')
                    tmc.update({
                        f'{tag}_peak_hour': self._xlsx_get_peak_hour(ws,position),
                        f'{tag}_north_bikes_vol': self._xlsx_get_north_bikes_vol(ws,position),
                        f'{tag}_north_peds_vol': self._xlsx_get_north_peds_vol(ws,position),
                        f'{tag}_north_veh_vol': self._xlsx_get_north_veh_vol(ws,position),
                        f'{tag}_east_bikes_vol' : self._xlsx_get_east_bikes_vol(ws,position),
                        f'{tag}_east_peds_vol' : self._xlsx_get_east_peds_vol(ws,position),
                        f'{tag}_east_veh_vol' : self._xlsx_get_east_veh_vol(ws,position),
                        f'{tag}_south_bikes_vol' : self._xlsx_get_south_bikes_vol(ws,position),
                        f'{tag}_south_peds_vol' : self._xlsx_get_south_peds_vol(ws,position),
                        f'{tag}_south_veh_vol' : self._xlsx_get_south_veh_vol(ws,position),
                        f'{tag}_west_bikes_vol' : self._xlsx_get_west_bikes_vol(ws,position),
                        f'{tag}_west_peds_vol' : self._xlsx_get_west_peds_vol(ws,position),
                        f'{tag}_west_veh_vol' : self._xlsx_get_west_veh_vol(ws,position)
                    })
            
            return tmc
        
        except Exception as e:
            logging.error(f'{url} had an error: {e}')
            return None

    def _xlsx_get_date(self, ws):
        element = ws['N15'].value
        return element if element else None

    def _xlsx_get_weather(self, ws):
        element = ws['V8'].value
        return element.lower() if element else None

    def _xlsx_get_type(self, ws):
        return None
    
    def _xlsx_get_peak_hour(self, ws, key):
        element_1 = ws[f'N{key+15}'].value
        element_2 = ws[f'P{key+15}'].value
        return str(element_1)[:-3].rjust(5, '0') + " - " + str(element_2)[:-3].rjust(5, '0') if element_1 and element_2 else None

    def _xlsx_get_north_bikes_vol(self, ws, key):
        element = ws[f'M{key+1}'].value
        return element if element else None

    def _xlsx_get_north_peds_vol(self, ws, key):
        element = ws[f'L{key+2}'].value
        return element if element else None

    def _xlsx_get_north_veh_vol(self, ws, key):
        element = ws[f'M{key}'].value
        return element if element else None
    
    def _xlsx_get_east_bikes_vol(self, ws, key):
        element = ws[f'AA{key+12}'].value
        return element if element else None
    
    def _xlsx_get_east_peds_vol(self, ws, key):
        element = ws[f'Z{key+11}'].value
        return element if element else None

    def _xlsx_get_east_veh_vol(self, ws, key):
        element = ws[f'AB{key+12}'].value
        return element if element else None
    
    def _xlsx_get_south_bikes_vol(self, ws, key):
        element = ws[f'P{key+26}'].value
        return element if element else None

    def _xlsx_get_south_peds_vol(self, ws, key):
        element = ws[f'Q{key+25}'].value
        return element if element else None
 
    def _xlsx_get_south_veh_vol(self, ws, key):
        element = ws[f'P{key+27}'].value
        return element if element else None
    
    def _xlsx_get_west_bikes_vol(self, ws, key):
        element = ws[f'B{key+15}'].value
        return element if element else None

    def _xlsx_get_west_peds_vol(self, ws, key):
        element = ws[f'C{key+16}'].value
        return element if element else None

    def _xlsx_get_west_veh_vol(self, ws, key):
        element = ws[f'A{key+15}'].value
        return element if element else None
    
    def _read_pdf(self,url):

        try:
            response = self.session.get(url)
            response.raise_for_status()
            pdf_stream = BytesIO(response.content)
            doc = fitz.open("pdf", pdf_stream.getvalue())
            text = doc[0].get_text()
            
            tmc = {}
            
            available_am = False
            available_md = False
            available_pm = False
            
            control_dict = {
                'AM':{'check':None,'table_key':None},
                'MD':{'check':None,'table_key':None},
                'PM':{'check':None,'table_key':None}
                }
            periods = re.findall(r"Maximum Hour\s*\n*\s*(\d{2}:\d{2} - \d{2}:\d{2})", text)
            for i in range(0,len(periods)):
                period = periods[i]
                if period:
                    peak_start_hr = int(period[:2])

                    if 6 <= peak_start_hr <= 8:
                        control_dict['AM']['table_key'] = i
                        available_am = True
                    elif 9 <= peak_start_hr <= 15:
                        control_dict['MD']['table_key'] = i
                        available_md = True
                    elif 16 <= peak_start_hr <= 19:
                        control_dict['PM']['table_key'] = i
                        available_pm = True
                    else:
                        continue
            
            control_dict['AM']['check'] = self.collect_am and available_am
            control_dict['MD']['check'] = self.collect_md and available_md
            control_dict['PM']['check'] = self.collect_pm and available_pm
            
            tmc.update({
                'date': self._get_date(text),
                'weather': self._get_weather(text),
                'type': self._get_type(text),
                'AM_available': available_am,
                'AM_scraped': self.collect_am,
                'MD_available': available_md,
                'MD_scraped': self.collect_md,                
                'PM_available': available_pm,
                'PM_scraped': self.collect_pm
            })
                
            for key,value in control_dict.items():
                if value.get('check'):
                    tag = key
                    position = value.get('table_key')
                    peds_and_bikes_vol = self._get_peds_and_bikes_vol(text,position)
                    veh_vol = self._get_veh_vol(text,position)
                    tmc.update({
                        f'{tag}_peak_hour': self._get_peak_hour(text,position),
                        f'{tag}_north_bikes_vol': peds_and_bikes_vol[0] if len(peds_and_bikes_vol) > 0 else None,
                        f'{tag}_north_peds_vol': peds_and_bikes_vol[1] if len(peds_and_bikes_vol) > 1 else None,
                        f'{tag}_north_veh_vol': veh_vol.get('north', None),
                        f'{tag}_east_bikes_vol' : peds_and_bikes_vol[5] if len(peds_and_bikes_vol) > 5 else None,
                        f'{tag}_east_peds_vol' : peds_and_bikes_vol[4] if len(peds_and_bikes_vol) > 4 else None,
                        f'{tag}_east_veh_vol' : veh_vol.get('east', None),
                        f'{tag}_south_bikes_vol' : peds_and_bikes_vol[7] if len(peds_and_bikes_vol) > 7 else None,
                        f'{tag}_south_peds_vol' : peds_and_bikes_vol[6] if len(peds_and_bikes_vol) > 6 else None,
                        f'{tag}_south_veh_vol' : veh_vol.get('south', None),
                        f'{tag}_west_bikes_vol' : peds_and_bikes_vol[2] if len(peds_and_bikes_vol) > 2 else None,
                        f'{tag}_west_peds_vol' : peds_and_bikes_vol[3] if len(peds_and_bikes_vol) > 3 else None,
                        f'{tag}_west_veh_vol' : veh_vol.get('west', None)
                    })
            
            return tmc
            
        except Exception as e:
            logging.error(f'{url} had an error: {e}')
            return None

    def _get_date(self, text):
        elements = re.findall(r"\b\w+,\s+\w+\s+\d{1,2},\s+\d{4}\b", text)
        return pd.to_datetime(elements[0] if elements else None)

    def _get_weather(self, text):
        elements = re.findall(r"Weather:\s+(\w+)", text)
        return elements[0].strip().lower() if elements else None

    def _get_type(self, text):
        element = re.findall(r"\((.*?)\)", text)
        return element[0] if element else None

    def _get_peak_hour(self, text, key):
        elements = re.findall(r"Maximum Hour\s*\n*\s*(\d{2}:\d{2} - \d{2}:\d{2})", text)
        return elements[key] if elements else None

    def _get_peds_and_bikes_vol(self, text, key):
        sections = re.split(r"Maximum Hour", text)[1:] 
        elements = []
        for section in sections:
            matches = re.findall(r"(Bikes|Peds|PEDs)\s*(\d+)", section)
            numbers = [number for keyword, number in matches]
            elements.append(numbers[:8])
            
        return elements[key] if elements else None

    def _get_veh_vol(self, text, key):
        
        sections = re.split(r"Maximum Hour", text, flags=re.DOTALL)
        patterns = [
            r"PEDs\s+\d+((?:\s+\d+)*).*?Bikes",  
            r"Peds\s+\d+((?:\s+\d+)*)\s+Peds",   
            r"Peds\s+\d+.*?Bikes\s+\d+((?:\s+\d+)*)\s+PEDs", 
        ]

        results = []
        elements = {}
        
        for pattern in patterns:
            match = re.search(pattern, sections[1:][key], re.DOTALL)
            if match:

                numbers = [int(num) for num in match.group(1).strip().split()]
                results.append(numbers)
            else:
                results.append([]) 
                
        if results[1]:
            if len(results[1]) > 4:
                elements.update({
                    'east' : results[1][-1], 
                    'west' : results[1][0]
                    })
                
            elif len(results[1]) == 4:
                
                if results[1][0] == results[1][1] and results[1][2] == results[1][3]:
                    elements.update({
                        'east' : results[1][-1], 
                        'west' : results[1][0]
                        })
                else:
                    decision = max(results[1])
                    elements.update({
                        'east' : decision if decision == results[1][-1] else None,
                        'west' : decision if decision == results[1][0] else None
                        })

            elif len(results[1]) == 3:
                decision = max(results[1])
                elements.update({
                    'east' : decision if decision == results[1][-1] else None,
                    'west' : decision if decision == results[1][0] else None
                    })
                
            elif len(results[1]) == 2:
                elements.update({
                    'east' : 'manual_check', 
                    'west' : 'manual_check'
                    })
                
            else:
                pass
            
        elements.update({
        'north' : results[0][0] if results[0] else None,
        'south' : max(results[2]) if results[2] else None
        })
        return elements
    
def main():
    # Setup for Recent Traffic Movement Counts (TMC)
    recent_tmc_raw_data_directory = os.path.join('..','data','raw','recent_intersection-traffic-movement-counts.csv')
    recent_tmc = RecentTMCScraper(raw_directory= recent_tmc_raw_data_directory ,period_filters=['AM','MD','PM'])
    recent_tmc_links = recent_tmc.scrape_sub_urls()
    recent_tmc_links.to_csv(os.path.join('..','data','scraped','recent_tmc_links.csv'),index=False)
    recent_tmc_scraped = recent_tmc.scrape_tmc_counts(recent_tmc_links)
    recent_tmc_scraped.insert(1,'intersection',recent_tmc_links['Intersection'])
    recent_tmc_scraped.to_csv(os.path.join('..','data','scraped','recent_tmc_scraped.csv'),index=False)

    # Setup for Old Traffic Movement Counts (TMC)
    old_tmc_raw_data_directory = os.path.join('..','data','raw','old_intersection-traffic-movement-counts.csv')
    old_tmc = OldTMCScraper(raw_directory= old_tmc_raw_data_directory ,period_filters=['AM','MD','PM'])
    old_tmc_links = old_tmc.scrape_sub_urls()
    old_tmc_links.to_csv(os.path.join('..','data','scraped','old_tmc_links.csv'),index=False)
    old_tmc_scraped = old_tmc.scrape_tmc_counts(old_tmc_links)
    old_tmc_scraped.insert(1,'intersection',old_tmc_links['INTERSECTION'])
    old_tmc_scraped.insert(2,'longitude',old_tmc_links['Geom'].apply(lambda x: re.findall(r'-\d+.\d+',x.split()[1])[0]))
    old_tmc_scraped.insert(3,'latitude',old_tmc_links['Geom'].apply(lambda x: re.findall(r'\d+.\d+',x.split()[2])[0]))
    old_tmc_scraped.to_csv(os.path.join('..','data','scraped','old_tmc_scraped.csv'),index=False)

if __name__ == "__main__":
    main()